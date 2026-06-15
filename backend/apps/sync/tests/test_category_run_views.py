"""Category-run report API: /api/sync/category-runs/ (list, detail, export)."""

import uuid
from datetime import timedelta
from io import BytesIO

import pytest
from django.utils import timezone

from apps.sites.tests.factories import HostingFactory, SiteFactory
from apps.sync.models import SyncLog
from apps.sync.services import CATEGORY_OPERATION


def _log(site, run_id, *, status=SyncLog.Status.SUCCESS, error="", **detail):
    defaults = {
        "site_name": site.name if site else "site-da-xoa.example",
        "site_url": site.base_url if site else "",
        "hosting": "TenTen",
    }
    return SyncLog.objects.create(
        site=site,
        operation=CATEGORY_OPERATION,
        status=status,
        run_id=run_id,
        error=error,
        detail={**defaults, **detail},
    )


@pytest.mark.django_db
def test_list_groups_rows_by_run(client):
    s1, s2 = SiteFactory(), SiteFactory()
    older, newer = uuid.uuid4(), uuid.uuid4()
    _log(s1, older, pulled=3, mapped=3)
    _log(s2, older, status=SyncLog.Status.ERROR, error="ConnectError")
    _log(s1, newer, pulled=5, mapped=4)
    # Legacy rows (pre-run_id) and other operations never appear.
    SyncLog.objects.create(
        site=s1, operation=CATEGORY_OPERATION, status=SyncLog.Status.SUCCESS, detail={}
    )
    SyncLog.objects.create(
        site=s1,
        operation="push_products",
        status=SyncLog.Status.SUCCESS,
        run_id=uuid.uuid4(),
        detail={},
    )

    resp = client.get("/api/sync/category-runs/")
    assert resp.status_code == 200
    rows = resp.data["results"]
    assert [r["run_id"] for r in rows] == [str(newer), str(older)]  # newest first

    newest, oldest = rows
    assert newest["site_count"] == 1
    assert newest["total_pulled"] == 5
    assert newest["total_mapped"] == 4
    assert newest["status"] == "success"
    # Mixed success+error rolls up to partial.
    assert oldest["site_count"] == 2
    assert oldest["total_pulled"] == 3
    assert oldest["error_count"] == 1
    assert oldest["status"] == "partial"


@pytest.mark.django_db
def test_list_paginates(client):
    site = SiteFactory()
    for _ in range(3):
        _log(site, uuid.uuid4(), pulled=1, mapped=1)

    resp = client.get("/api/sync/category-runs/", {"page_size": 2})
    assert resp.status_code == 200
    assert resp.data["count"] == 3
    assert len(resp.data["results"]) == 2


@pytest.mark.django_db
def test_detail_returns_per_site_rows_with_categories(client):
    hosting = HostingFactory(provider="TenTen")
    site = SiteFactory(name="A-Site", hosting=hosting)
    run = uuid.uuid4()
    cats = [
        {"woo_id": 10, "woo_name": "Pin mặt trời", "hub_id": 1, "hub_name": "Pin mặt trời"},
        {"woo_id": 11, "woo_name": " Pin  mặt trời ", "hub_id": 1, "hub_name": "Pin mặt trời"},
    ]
    _log(site, run, pulled=2, mapped=1, categories=cats)
    _log(None, run, status=SyncLog.Status.ERROR, error="ConnectError")  # site deleted since

    resp = client.get(f"/api/sync/category-runs/{run}/")
    assert resp.status_code == 200
    assert resp.data["run_id"] == str(run)
    assert resp.data["status"] == "partial"
    assert resp.data["site_count"] == 2

    live, gone = resp.data["sites"]
    assert live["site_name"] == "A-Site"
    assert live["hosting"] == "TenTen"  # live FK wins
    assert [c["woo_id"] for c in live["categories"]] == [10, 11]
    assert live["categories"][1]["hub_name"] == "Pin mặt trời"
    # Deleted site falls back to the snapshot taken at pull time.
    assert gone["site_id"] is None
    assert gone["site_name"] == "site-da-xoa.example"
    assert gone["hosting"] == "TenTen"
    assert gone["error"] == "ConnectError"
    assert gone["categories"] == []


@pytest.mark.django_db
def test_cleared_runs_are_excluded_from_report(client):
    """After "clear category sync data" soft-deletes the rows, the run list,
    stats and detail all behave as if the history is empty."""
    site = SiteFactory()
    run = uuid.uuid4()
    _log(site, run, pulled=3, mapped=3)

    SyncLog.objects.filter(operation=CATEGORY_OPERATION).update(is_deleted=True)

    assert client.get("/api/sync/category-runs/").data["count"] == 0
    assert client.get("/api/sync/category-runs/stats/").data["total"] == 0
    assert client.get(f"/api/sync/category-runs/{run}/").status_code == 404


# --- duration / triggered_by / site_label + stats + filters -------------------


@pytest.mark.django_db
def test_list_includes_duration_user_and_site_label(client, django_user_model):
    user = django_user_model.objects.create_user(
        username="admin", first_name="Nguyễn", last_name="Văn A", password="x"
    )
    site = SiteFactory(name="solarcity.com.vn")
    run = uuid.uuid4()
    start = timezone.now()
    log = SyncLog.objects.create(
        site=site,
        operation=CATEGORY_OPERATION,
        status=SyncLog.Status.SUCCESS,
        run_id=run,
        triggered_by=user,
        started_at=start,
        detail={"pulled": 5, "mapped": 4, "site_name": site.name},
    )
    # auto_now_add stamps created_at at insert; push it 90s out to fake a finish.
    SyncLog.objects.filter(id=log.id).update(created_at=start + timedelta(seconds=90))

    row = client.get("/api/sync/category-runs/").data["results"][0]
    assert row["site_label"] == "solarcity.com.vn"  # single-site run → named
    assert row["triggered_by"] == "Nguyễn Văn A"  # full name wins over username
    assert row["duration_seconds"] == 90


@pytest.mark.django_db
def test_list_multi_site_run_has_no_site_label(client):
    s1, s2 = SiteFactory(), SiteFactory()
    run = uuid.uuid4()
    _log(s1, run, pulled=1, mapped=1)
    _log(s2, run, pulled=1, mapped=1)

    row = client.get("/api/sync/category-runs/").data["results"][0]
    assert row["site_count"] == 2
    assert row["site_label"] is None  # multi-site → UI shows "N site"
    assert row["triggered_by"] is None  # _log records no user


@pytest.mark.django_db
def test_list_filter_by_status(client):
    site = SiteFactory()
    ok, bad = uuid.uuid4(), uuid.uuid4()
    _log(site, ok, pulled=1, mapped=1)
    _log(site, bad, status=SyncLog.Status.ERROR, error="ConnectError")

    def run_ids(params):
        return {r["run_id"] for r in client.get("/api/sync/category-runs/", params).data["results"]}

    assert run_ids({"status": "success"}) == {str(ok)}
    assert run_ids({"status": "error"}) == {str(bad)}


@pytest.mark.django_db
def test_list_filter_by_site(client):
    s1, s2 = SiteFactory(name="alpha.vn"), SiteFactory(name="beta.vn")
    r1, r2 = uuid.uuid4(), uuid.uuid4()
    _log(s1, r1, pulled=1, mapped=1)
    _log(s2, r2, pulled=1, mapped=1)

    rows = client.get("/api/sync/category-runs/", {"site": s1.id}).data["results"]
    assert {r["run_id"] for r in rows} == {str(r1)}


@pytest.mark.django_db
def test_list_search_by_site_name(client):
    s1, s2 = SiteFactory(name="alpha.vn"), SiteFactory(name="beta.vn")
    r1, r2 = uuid.uuid4(), uuid.uuid4()
    _log(s1, r1, pulled=1, mapped=1)
    _log(s2, r2, pulled=1, mapped=1)

    rows = client.get("/api/sync/category-runs/", {"search": "alpha"}).data["results"]
    assert {r["run_id"] for r in rows} == {str(r1)}


@pytest.mark.django_db
def test_stats_counts_split_and_last_run(client):
    site = SiteFactory(name="solarcity.com.vn")
    ok, bad, mixed = uuid.uuid4(), uuid.uuid4(), uuid.uuid4()
    _log(site, ok, pulled=2, mapped=2)
    _log(site, bad, status=SyncLog.Status.ERROR, error="ConnectError")
    _log(site, mixed, pulled=1, mapped=1)  # created last → newest
    _log(SiteFactory(), mixed, status=SyncLog.Status.ERROR, error="X")

    data = client.get("/api/sync/category-runs/stats/").data
    assert data["total"] == 3
    assert data["success"] == 1
    assert data["error"] == 1
    assert data["partial"] == 1
    assert data["last_run"]["run_id"] == str(mixed)
    assert data["last_run"]["status"] == "partial"
    assert data["last_run"]["site_count"] == 2


@pytest.mark.django_db
def test_stats_default_window_excludes_old_runs(client):
    site = SiteFactory()
    old, recent = uuid.uuid4(), uuid.uuid4()
    old_log = _log(site, old, pulled=1, mapped=1)
    SyncLog.objects.filter(id=old_log.id).update(
        created_at=timezone.now() - timedelta(days=45)
    )
    _log(site, recent, pulled=1, mapped=1)

    data = client.get("/api/sync/category-runs/stats/").data
    assert data["total"] == 1  # 45-day-old run is outside the default 30-day window


@pytest.mark.django_db
def test_detail_unknown_run_404(client):
    resp = client.get(f"/api/sync/category-runs/{uuid.uuid4()}/")
    assert resp.status_code == 404


@pytest.mark.django_db
def test_export_returns_readable_xlsx(client):
    from openpyxl import load_workbook

    site = SiteFactory(name="A-Site", hosting=HostingFactory(provider="TenTen"))
    run = uuid.uuid4()
    _log(
        site,
        run,
        pulled=2,
        mapped=2,
        categories=[
            {"woo_id": 10, "woo_name": "Pin mặt trời", "hub_id": 1, "hub_name": "Pin mặt trời"},
            {"woo_id": 12, "woo_name": "Inverter", "hub_id": 2, "hub_name": "Inverter"},
        ],
    )

    resp = client.get(f"/api/sync/category-runs/{run}/export/")
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="bao-cao-danh-muc-' in resp["Content-Disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["Tổng quan", "Chi tiết"]

    overview = wb["Tổng quan"]
    # Row 5 = table header, row 6 = the one site row.
    assert overview.cell(row=6, column=1).value == "A-Site"
    assert overview.cell(row=6, column=3).value == "TenTen"
    assert overview.cell(row=6, column=5).value == 2  # pulled

    detail = wb["Chi tiết"]
    # Row 1 = timestamp, row 2 = header, rows 3-4 = the categories.
    assert detail.cell(row=3, column=4).value == "Pin mặt trời"
    assert detail.cell(row=4, column=4).value == "Inverter"
    assert detail.cell(row=4, column=6).value == "Inverter"  # hub name


@pytest.mark.django_db
def test_export_unknown_run_404(client):
    resp = client.get(f"/api/sync/category-runs/{uuid.uuid4()}/export/")
    assert resp.status_code == 404
