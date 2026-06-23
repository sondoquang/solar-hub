"""Product-push run report API: /api/sync/product-runs/ (list, detail, stats, export)."""

import uuid
from datetime import timedelta
from io import BytesIO

import pytest
from django.utils import timezone

from apps.sites.tests.factories import HostingFactory, SiteFactory
from apps.sync.models import SyncLog
from apps.sync.services import PRODUCT_OPERATION


def _plog(
    site,
    run_id,
    *,
    status=SyncLog.Status.SUCCESS,
    error="",
    created=0,
    updated=0,
    deleted=0,
    **detail,
):
    base = {
        "site_name": site.name if site else "site-da-xoa.example",
        "site_url": site.base_url if site else "",
        "hosting": "TenTen",
    }
    return SyncLog.objects.create(
        site=site,
        operation=PRODUCT_OPERATION,
        status=status,
        run_id=run_id,
        error=error,
        created_count=created,
        updated_count=updated,
        deleted_count=deleted,
        detail={**base, **detail},
    )


@pytest.mark.django_db
def test_list_groups_rows_by_run(client):
    s1, s2 = SiteFactory(), SiteFactory()
    older, newer = uuid.uuid4(), uuid.uuid4()
    _plog(s1, older, created=3, updated=1, adopted_count=2)
    _plog(s2, older, status=SyncLog.Status.ERROR, error="ConnectError")
    _plog(s1, newer, created=5, updated=2, adopted_count=1)
    # Legacy rows (pre-run_id) and other operations never appear.
    SyncLog.objects.create(
        site=s1, operation=PRODUCT_OPERATION, status=SyncLog.Status.SUCCESS, detail={}
    )
    SyncLog.objects.create(
        site=s1,
        operation="pull_categories",
        status=SyncLog.Status.SUCCESS,
        run_id=uuid.uuid4(),
        detail={},
    )

    resp = client.get("/api/sync/product-runs/")
    assert resp.status_code == 200
    rows = resp.data["results"]
    assert [r["run_id"] for r in rows] == [str(newer), str(older)]  # newest first

    newest, oldest = rows
    assert newest["site_count"] == 1
    assert newest["total_created"] == 5
    assert newest["total_updated"] == 2
    assert newest["total_adopted"] == 1
    assert newest["status"] == "success"
    # Mixed success+error rolls up to partial.
    assert oldest["site_count"] == 2
    assert oldest["total_created"] == 3
    assert oldest["total_adopted"] == 2
    assert oldest["error_count"] == 1
    assert oldest["status"] == "partial"


@pytest.mark.django_db
def test_detail_returns_per_site_rows_with_failures(client):
    hosting = HostingFactory(provider="TenTen")
    site = SiteFactory(name="A-Site", hosting=hosting)
    run = uuid.uuid4()
    failed = [
        {"sku": "SP-1", "op": "create", "code": "product_invalid_sku", "message": "dup"},
        {"sku": "SP-2", "op": "update", "code": "rest_cannot_update", "message": "boom"},
    ]
    _plog(
        site,
        run,
        status=SyncLog.Status.PARTIAL,
        created=4,
        updated=1,
        adopted_count=3,
        adopted=["SP-A"],
        failed=failed,
    )
    _plog(None, run, status=SyncLog.Status.ERROR, error="ConnectError")  # site deleted since

    resp = client.get(f"/api/sync/product-runs/{run}/")
    assert resp.status_code == 200
    assert resp.data["run_id"] == str(run)
    assert resp.data["status"] == "partial"
    assert resp.data["site_count"] == 2
    assert resp.data["total_adopted"] == 3
    assert resp.data["total_failed"] == 2

    live, gone = resp.data["sites"]
    assert live["site_name"] == "A-Site"
    assert live["hosting"] == "TenTen"  # live FK wins
    assert live["adopted_count"] == 3
    # kind classifies duplicate-SKU vs genuine error for UI tinting.
    kinds = {f["sku"]: f["kind"] for f in live["failed"]}
    assert kinds == {"SP-1": "duplicate", "SP-2": "error"}
    # Deleted site falls back to the snapshot taken at push time.
    assert gone["site_id"] is None
    assert gone["site_name"] == "site-da-xoa.example"
    assert gone["error"] == "ConnectError"
    assert gone["failed"] == []


@pytest.mark.django_db
def test_list_filter_by_status_and_site(client):
    s1, s2 = SiteFactory(name="alpha.vn"), SiteFactory(name="beta.vn")
    ok, bad = uuid.uuid4(), uuid.uuid4()
    _plog(s1, ok, created=1)
    _plog(s2, bad, status=SyncLog.Status.ERROR, error="ConnectError")

    def run_ids(params):
        return {r["run_id"] for r in client.get("/api/sync/product-runs/", params).data["results"]}

    assert run_ids({"status": "success"}) == {str(ok)}
    assert run_ids({"status": "error"}) == {str(bad)}
    assert run_ids({"site": s1.id}) == {str(ok)}
    assert run_ids({"search": "alpha"}) == {str(ok)}


@pytest.mark.django_db
def test_stats_counts_split_and_last_run(client):
    site = SiteFactory(name="solarcity.com.vn")
    ok, bad, mixed = uuid.uuid4(), uuid.uuid4(), uuid.uuid4()
    _plog(site, ok, created=2)
    _plog(site, bad, status=SyncLog.Status.ERROR, error="ConnectError")
    _plog(site, mixed, created=1)  # created last → newest
    _plog(SiteFactory(), mixed, status=SyncLog.Status.ERROR, error="X")

    data = client.get("/api/sync/product-runs/stats/").data
    assert data["total"] == 3
    assert data["success"] == 1
    assert data["error"] == 1
    assert data["partial"] == 1
    assert data["last_run"]["run_id"] == str(mixed)
    assert data["last_run"]["status"] == "partial"


@pytest.mark.django_db
def test_stats_default_window_excludes_old_runs(client):
    site = SiteFactory()
    old, recent = uuid.uuid4(), uuid.uuid4()
    old_log = _plog(site, old, created=1)
    SyncLog.objects.filter(id=old_log.id).update(created_at=timezone.now() - timedelta(days=45))
    _plog(site, recent, created=1)

    assert client.get("/api/sync/product-runs/stats/").data["total"] == 1


@pytest.mark.django_db
def test_detail_unknown_run_404(client):
    assert client.get(f"/api/sync/product-runs/{uuid.uuid4()}/").status_code == 404


@pytest.mark.django_db
def test_export_returns_readable_xlsx(client):
    from openpyxl import load_workbook

    site = SiteFactory(name="A-Site", hosting=HostingFactory(provider="TenTen"))
    run = uuid.uuid4()
    _plog(
        site,
        run,
        status=SyncLog.Status.PARTIAL,
        created=2,
        updated=1,
        adopted_count=1,
        failed=[{"sku": "SP-1", "op": "create", "code": "product_invalid_sku", "message": "dup"}],
    )

    resp = client.get(f"/api/sync/product-runs/{run}/export/")
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="bao-cao-san-pham-' in resp["Content-Disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["Tổng quan", "Chi tiết"]

    overview = wb["Tổng quan"]
    # Row 5 = table header, row 6 = the one site row.
    assert overview.cell(row=6, column=1).value == "A-Site"
    assert overview.cell(row=6, column=3).value == "TenTen"
    assert overview.cell(row=6, column=5).value == 2  # created

    detail = wb["Chi tiết"]
    # Row 1 = timestamp, row 2 = header, row 3 = the one failed SKU.
    assert detail.cell(row=3, column=3).value == "SP-1"
    assert detail.cell(row=3, column=5).value == "duplicate"  # kind
    assert detail.cell(row=3, column=6).value == "product_invalid_sku"


@pytest.mark.django_db
def test_export_unknown_run_404(client):
    assert client.get(f"/api/sync/product-runs/{uuid.uuid4()}/export/").status_code == 404
