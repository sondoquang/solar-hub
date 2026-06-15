"""Category-run report: group ``pull_categories`` SyncLog rows by ``run_id``.

One "run" = one user click of "sync categories" (one fan-out across sites).
Each site writes one SyncLog row stamped with the run_id and carrying a
snapshot in ``detail`` (site name/url/hosting + the per-category woo→hub list,
see ``apps.catalog.services.pull_categories_for_site``). This module rolls
those rows up for the report endpoints and builds the Excel export.
"""

import uuid
from datetime import timedelta
from io import BytesIO

from django.db.models import Count, F, Min, Q
from django.utils import timezone

from .models import SyncLog

# Same value as apps.catalog.services.CATEGORY_OPERATION; redeclared to avoid
# importing the (heavy, httpx-bound) catalog service module at startup.
CATEGORY_OPERATION = "pull_categories"


def category_runs_queryset(params=None):
    """Grouped queryset of runs (newest first), one row per ``run_id``.

    Legacy rows written before run_id existed are excluded — they carry no
    category snapshot, so there is nothing to report. ``params`` (the request
    query params) optionally filter the list:

    - ``date_from`` / ``date_to`` — calendar range on the row timestamps.
    - ``site`` — runs that touched that site id (full run still shown).
    - ``status`` — rolled-up result (``success``/``partial``/``error``): a run is
      success when every row succeeded, error when every row errored, else
      partial — expressed against the per-run success/error counts.
    - ``search`` — matches a site name, the triggering user, or an exact run id;
      a run shows up if ANY of its rows matches.

    Each row is annotated with ``n_total``/``n_success``/``n_error`` so the
    status filter (and the stats roll-up) need no extra query.
    """
    params = params or {}
    logs = SyncLog.objects.filter(
        operation=CATEGORY_OPERATION, run_id__isnull=False, is_deleted=False
    )

    date_from = params.get("date_from")
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    date_to = params.get("date_to")
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    site = params.get("site")
    if site and str(site).isdigit():
        run_ids_for_site = SyncLog.objects.filter(
            operation=CATEGORY_OPERATION, site_id=int(site), is_deleted=False
        ).values("run_id")
        logs = logs.filter(run_id__in=run_ids_for_site)

    search = (params.get("search") or "").strip()
    if search:
        match = (
            Q(detail__site_name__icontains=search)
            | Q(site__name__icontains=search)
            | Q(triggered_by__username__icontains=search)
            | Q(triggered_by__first_name__icontains=search)
            | Q(triggered_by__last_name__icontains=search)
        )
        try:
            match |= Q(run_id=uuid.UUID(search))
        except (ValueError, AttributeError, TypeError):
            pass
        run_ids_matched = (
            SyncLog.objects.filter(operation=CATEGORY_OPERATION, is_deleted=False)
            .filter(match)
            .values("run_id")
        )
        logs = logs.filter(run_id__in=run_ids_matched)

    runs = (
        logs.values("run_id")
        .annotate(
            started_at=Min("created_at"),
            n_total=Count("id"),
            n_success=Count("id", filter=Q(status=SyncLog.Status.SUCCESS)),
            n_error=Count("id", filter=Q(status=SyncLog.Status.ERROR)),
        )
        .order_by("-started_at")
    )

    status = params.get("status")
    if status == SyncLog.Status.SUCCESS:
        runs = runs.filter(n_success=F("n_total"))
    elif status == SyncLog.Status.ERROR:
        runs = runs.filter(n_error=F("n_total"))
    elif status == SyncLog.Status.PARTIAL:
        runs = runs.exclude(n_success=F("n_total")).exclude(n_error=F("n_total"))
    return runs


def _rollup_status(logs) -> str:
    """All success → success, all error → error, mixed → partial."""
    statuses = {log.status for log in logs}
    if statuses == {SyncLog.Status.SUCCESS}:
        return SyncLog.Status.SUCCESS
    if statuses == {SyncLog.Status.ERROR}:
        return SyncLog.Status.ERROR
    return SyncLog.Status.PARTIAL


def _user_label(user) -> str:
    """Display name for the "Người chạy" column: full name, else username."""
    return user.get_full_name() or user.get_username()


def _summary(run_id, started_at, logs) -> dict:
    """Roll the run's per-site rows up into one report row.

    ``duration_seconds`` spans the earliest ``started_at`` (when the first site's
    pull began) to the latest ``created_at`` (when the last site finished) —
    falling back to ``started_at`` for legacy rows that never recorded a start.
    ``triggered_by`` is the first row's clicking user (one click = one user);
    ``site_label`` names the single site for one-site runs, else null (the
    frontend shows "N site")."""
    finished = max((log.created_at for log in logs), default=started_at)
    starts = [log.started_at for log in logs if log.started_at] or [started_at]
    duration = (finished - min(starts)).total_seconds() if logs else 0

    triggered_by = next(
        (_user_label(log.triggered_by) for log in logs if log.triggered_by_id), None
    )
    site_label = None
    if len(logs) == 1:
        log = logs[0]
        site_label = log.site.name if log.site else (log.detail or {}).get("site_name", "")

    return {
        "run_id": str(run_id),
        "started_at": started_at,
        "site_count": len(logs),
        "total_pulled": sum(int((log.detail or {}).get("pulled") or 0) for log in logs),
        "total_mapped": sum(int((log.detail or {}).get("mapped") or 0) for log in logs),
        "error_count": sum(1 for log in logs if log.status == SyncLog.Status.ERROR),
        "status": _rollup_status(logs),
        "duration_seconds": max(0, round(duration)),
        "triggered_by": triggered_by,
        "site_label": site_label,
    }


def summarize_runs(run_rows) -> list[dict]:
    """Roll the page's runs up in Python — bounded work (page_size × sites/run),
    and JSON-key aggregates in SQL would be fragile for the ``detail`` counts."""
    run_ids = [row["run_id"] for row in run_rows]
    logs_by_run: dict = {}
    for log in SyncLog.objects.filter(
        operation=CATEGORY_OPERATION, run_id__in=run_ids, is_deleted=False
    ).select_related("site", "triggered_by"):
        logs_by_run.setdefault(log.run_id, []).append(log)
    return [
        _summary(row["run_id"], row["started_at"], logs_by_run.get(row["run_id"], []))
        for row in run_rows
    ]


def category_run_stats(params=None) -> dict:
    """Stat-card counts for the Lịch sử đồng bộ tab over a window (default last
    30 days when no explicit date range is given): total runs and the
    success/partial/error split, plus the most-recent run as ``last_run``.

    Reuses ``category_runs_queryset`` (so site/search filters apply identically),
    minus the ``status`` filter — the split is what we're counting."""
    params = dict(params or {})
    if not params.get("date_from") and not params.get("date_to"):
        params["date_from"] = (timezone.localdate() - timedelta(days=30)).isoformat()

    runs = list(
        category_runs_queryset(
            {
                "site": params.get("site"),
                "date_from": params.get("date_from"),
                "date_to": params.get("date_to"),
                "search": params.get("search"),
            }
        ).values("run_id", "started_at", "n_total", "n_success", "n_error")
    )

    success = sum(1 for r in runs if r["n_success"] == r["n_total"])
    error = sum(1 for r in runs if r["n_error"] == r["n_total"])

    last_run = None
    if runs:
        top = runs[0]  # queryset is ordered -started_at
        logs = list(
            SyncLog.objects.filter(
                operation=CATEGORY_OPERATION, run_id=top["run_id"], is_deleted=False
            ).select_related("site", "triggered_by")
        )
        summ = _summary(top["run_id"], top["started_at"], logs)
        last_run = {
            "run_id": summ["run_id"],
            "started_at": summ["started_at"],
            "site_count": summ["site_count"],
            "site_label": summ["site_label"],
            "status": summ["status"],
        }

    return {
        "total": len(runs),
        "success": success,
        "partial": len(runs) - success - error,
        "error": error,
        "last_run": last_run,
    }


def _site_row(log) -> dict:
    """One per-site row, null-safe: ``site`` is SET_NULL on delete, so fall
    back to the snapshot taken into ``detail`` at pull time."""
    site = log.site
    detail = log.detail or {}
    hosting = site.hosting if site else None
    return {
        "site_id": site.id if site else None,
        "site_name": site.name if site else detail.get("site_name", ""),
        "site_url": site.base_url if site else detail.get("site_url", ""),
        "hosting": (
            (hosting.provider or hosting.name)
            if hosting
            else detail.get("hosting", "")
        ),
        "status": log.status,
        "error": log.error,
        "pulled": int(detail.get("pulled") or 0),
        "mapped": int(detail.get("mapped") or 0),
        "categories": detail.get("categories") or [],
        "created_at": log.created_at,
    }


# Operations that fan out one SyncLog row per site and can be polled live by the
# progress banner. Each manual "Đồng bộ ngay"/"Đồng bộ danh mục" click stamps its
# rows with a fresh ``run_id``; the banner counts how many have landed.
PROGRESS_OPERATIONS = {
    CATEGORY_OPERATION,  # pull_categories
    "push_products",
    "poll_orders",
}


def run_progress(run_id, operation) -> dict:
    """Live progress of one fan-out run: how many of its per-site ``SyncLog``
    rows have landed so far (and how many errored), for the polling progress
    banner. Generic over ``operation`` so orders/products/categories share it.

    Returns ``done=0`` (never 404) before the first site finishes — the banner
    polls a freshly-triggered run and just waits for ``done`` to climb to the
    ``expected`` count the trigger endpoint returned."""
    rows = list(
        SyncLog.objects.filter(
            operation=operation, run_id=run_id, is_deleted=False
        ).values_list("status", flat=True)
    )
    return {
        "run_id": str(run_id),
        "operation": operation,
        "done": len(rows),
        "error_count": sum(1 for s in rows if s == SyncLog.Status.ERROR),
    }


def run_detail(run_id) -> dict | None:
    """Summary + per-site rows (incl. category snapshots) of one run, or None."""
    logs = list(
        SyncLog.objects.filter(
            operation=CATEGORY_OPERATION, run_id=run_id, is_deleted=False
        )
        .select_related("site__hosting", "triggered_by")
        .order_by("created_at")
    )
    if not logs:
        return None
    summary = _summary(run_id, min(log.created_at for log in logs), logs)
    summary["sites"] = [_site_row(log) for log in logs]
    return summary


# --- Excel export -------------------------------------------------------------

_OVERVIEW_HEADERS = [
    "Site",
    "URL",
    "Hosting",
    "Trạng thái",
    "Số danh mục (Woo)",
    "Đã ánh xạ (Hub)",
    "Lỗi",
]
_DETAIL_HEADERS = [
    "Site",
    "Hosting",
    "Woo ID",
    "Tên danh mục (Woo)",
    "Hub ID",
    "Tên danh mục (Hub)",
]


def build_run_workbook(detail: dict) -> bytes:
    """Two-sheet .xlsx for one run: "Tổng quan" (one row per site) and a flat
    "Chi tiết" (one row per category per site — flat instead of per-site sheets
    to dodge the 31-char sheet-name limit and duplicate site names). In-memory
    via BytesIO: a run is at most a few thousand rows, same class of inline work
    as the existing CSV/PDF exports.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    started = timezone.localtime(detail["started_at"]).strftime("%d/%m/%Y %H:%M:%S")

    wb = Workbook()
    overview = wb.active
    overview.title = "Tổng quan"
    overview.append(["Báo cáo đồng bộ danh mục"])
    overview["A1"].font = bold
    overview.append(["Thời gian đồng bộ", started])
    overview.append(["Mã lần đồng bộ", detail["run_id"]])
    overview.append([])
    overview.append(_OVERVIEW_HEADERS)
    for cell in overview[overview.max_row]:
        cell.font = bold
    for row in detail["sites"]:
        overview.append(
            [
                row["site_name"],
                row["site_url"],
                row["hosting"],
                row["status"],
                row["pulled"],
                row["mapped"],
                row["error"],
            ]
        )

    sheet = wb.create_sheet("Chi tiết")
    sheet.append([f"Thời gian đồng bộ: {started}"])
    sheet["A1"].font = bold
    sheet.append(_DETAIL_HEADERS)
    for cell in sheet[2]:
        cell.font = bold
    for row in detail["sites"]:
        for cat in row["categories"]:
            sheet.append(
                [
                    row["site_name"],
                    row["hosting"],
                    cat.get("woo_id"),
                    cat.get("woo_name"),
                    cat.get("hub_id"),
                    cat.get("hub_name"),
                ]
            )

    widths = {overview: [28, 36, 18, 12, 18, 16, 24], sheet: [28, 18, 10, 40, 10, 40]}
    for ws, cols in widths.items():
        for i, width in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_export_filename(detail: dict) -> str:
    started = timezone.localtime(detail["started_at"]).strftime("%Y%m%d-%H%M%S")
    return f"bao-cao-danh-muc-{started}.xlsx"
