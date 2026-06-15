"""Service layer for the sites app: view → service → model/platform client.

Handles secret encryption and the on-demand connection test. Views/serializers
stay thin; all remote traffic goes through the platform client built by
``client_for_site`` (``WooClient`` or ``SapoClient``).
"""

import logging
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
from functools import partial

import httpx
import nh3
import openpyxl
from django.conf import settings
from django.db import connection, transaction
from django.db.models import Q
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from apps.integrations.sapo import SapoClient
from apps.integrations.woocommerce import WooClient

from .crypto import decrypt_secret, encrypt_secret
from .models import Hosting, Site, SiteNote, SiteNoteImage

logger = logging.getLogger(__name__)

IMPORT_COLUMNS = ["name", "base_url", "consumer_key", "consumer_secret"]

# Hosting import: only ``name`` is required; the rest are optional.
HOSTING_IMPORT_COLUMNS = ["name", "provider", "account_username", "note", "check_concurrency"]
HOSTING_REQUIRED_COLUMNS = ["name"]

DEFAULT_CHECK_CONCURRENCY = 5
MAX_CHECK_CONCURRENCY = 50


def create_site(
    *,
    name: str,
    base_url: str,
    consumer_key: str,
    consumer_secret: str,
    hosting=None,
    is_primary: bool = False,
    platform: str = Site.Platform.WOOCOMMERCE,
) -> Site:
    return Site.objects.create(
        name=name,
        base_url=base_url,
        consumer_key=consumer_key,
        consumer_secret_enc=encrypt_secret(consumer_secret),
        hosting=hosting,
        is_primary=is_primary,
        platform=platform,
    )


def delete_site(site: Site) -> None:
    site.is_deleted = True
    site.deleted_at = timezone.now()
    site.save(update_fields=["is_deleted", "deleted_at", "updated_at"])


def update_site(site: Site, *, consumer_secret: str | None = None, **fields) -> Site:
    for attr, value in fields.items():
        setattr(site, attr, value)
    if consumer_secret:
        site.consumer_secret_enc = encrypt_secret(consumer_secret)
    site.save()
    return site


def client_for_site(site: Site):
    """Build the platform client with the decrypted secret (in memory only).

    Dispatches on ``Site.platform``: WooCommerce sites get a ``WooClient``,
    Sapo Web sites a ``SapoClient`` (same method surface — see
    apps/integrations/sapo.py)."""
    secret = decrypt_secret(site.consumer_secret_enc)
    if site.platform == Site.Platform.SAPO:
        # Talk to the canonical ``*.mysapo.net`` host directly once discovered
        # (health-check persists it as ``sapo_store_host``). The storefront domain
        # only redirects the order LIST to the canonical host — per-order admin
        # paths (``/orders/{id}.json``, ``.../cancel.json``, ``.../transactions.json``)
        # bounce to a logout/login flow there and lose auth (→ 502). Fall back to
        # ``base_url`` + redirect-following for a site not yet health-checked.
        base_url = (
            f"https://{site.sapo_store_host}" if site.sapo_store_host else site.base_url
        )
        return SapoClient(
            base_url=base_url,
            api_key=site.consumer_key,
            api_secret=secret,
            status_timeout=settings.SITE_HEALTHCHECK_TIMEOUT_SECONDS,
            throttle_seconds=settings.SAPO_THROTTLE_SECONDS,
            max_429_retries=settings.SAPO_MAX_429_RETRIES,
            retry_after_default=settings.SAPO_RETRY_AFTER_DEFAULT_SECONDS,
        )
    return WooClient(
        base_url=site.base_url,
        consumer_key=site.consumer_key,
        consumer_secret=secret,
        status_timeout=settings.SITE_HEALTHCHECK_TIMEOUT_SECONDS,
    )


def test_connection(site: Site, *, check_type: str = "manual", performed_by=None) -> dict:
    """Call system_status to verify the key, timing the round-trip.

    Updates ``Site.status`` (up/down) and appends a ``HealthCheck`` history row
    via the monitoring app (status there is the 3-level health derived from the
    response time). ``check_type``/``performed_by`` flow through to that row so
    the history shows who ran it ("Hệ thống" for the periodic task).
    """
    ok = False
    detail = ""
    resolved_host = None
    started = time.monotonic()
    try:
        client = client_for_site(site)
        client.system_status()
        ok = True
        detail = "Kết nối thành công."
        # SapoClient records the canonical store host it landed on (after the
        # *.mysapo.net redirect); persist it so the order poll can dedupe
        # storefront domains sharing one store. WooClient has no such attribute.
        resolved_host = getattr(client, "resolved_host", None)
    except httpx.HTTPStatusError as exc:
        # The site answered with a non-2xx — capture status + a body snippet so
        # the cause is visible in the UI (HealthCheckDetailModal / test toast),
        # not just the exception class. Body is sliced on the decoded str
        # (UTF-8 safe); prefix + 180 chars stays under detail's 255 limit.
        status = exc.response.status_code
        body = (exc.response.text or "")[:180]
        detail = f"Lỗi HTTP {status}: {body}"
        logger.error(
            "test_connection failed site_id=%s status=%s body=%s", site.id, status, body
        )
    except httpx.HTTPError as exc:
        # Connect/timeout errors have no response to read.
        detail = f"Lỗi kết nối: {exc.__class__.__name__}"
        logger.error("test_connection failed site_id=%s: %s", site.id, exc)
    except Exception as exc:  # noqa: BLE001 — surface as down, but log the cause
        detail = f"Lỗi không xác định: {exc.__class__.__name__}"
        logger.error("test_connection error site_id=%s: %s", site.id, exc)
    response_time_ms = int((time.monotonic() - started) * 1000)

    site.status = Site.Status.UP if ok else Site.Status.DOWN
    site.last_checked_at = timezone.now()
    update_fields = ["status", "last_checked_at", "updated_at"]
    if resolved_host and site.sapo_store_host != resolved_host:
        site.sapo_store_host = resolved_host
        update_fields.append("sapo_store_host")
    site.save(update_fields=update_fields)

    # Lazy import: apps.monitoring.tasks imports apps.sites, so importing
    # monitoring at module load would be circular.
    from apps.monitoring import services as monitoring

    health = monitoring.record_check(
        site=site,
        ok=ok,
        response_time_ms=response_time_ms,
        detail=detail,
        check_type=check_type,
        performed_by=performed_by,
    )
    return {
        "ok": ok,
        "status": site.status,
        "detail": detail,
        "response_time_ms": response_time_ms,
        "health_status": health.status,
    }


def bulk_test_connections(sites, *, performed_by=None) -> list[dict]:
    """Test a list of sites sequentially (one at a time = natural throttle)."""
    return [
        {"id": s.id, **test_connection(s, performed_by=performed_by)}
        for s in sites
    ]


# --- Hosting -----------------------------------------------------------------


def create_hosting(*, name: str, **fields) -> Hosting:
    return Hosting.objects.create(name=name, **fields)


def update_hosting(hosting: Hosting, **fields) -> Hosting:
    for attr, value in fields.items():
        setattr(hosting, attr, value)
    hosting.save()
    return hosting


def delete_hosting(hosting: Hosting) -> None:
    """Soft-delete a hosting. Member sites keep their FK (the hosting is just
    hidden), so re-grouping is reversible by un-deleting it."""
    hosting.is_deleted = True
    hosting.deleted_at = timezone.now()
    hosting.save(update_fields=["is_deleted", "deleted_at", "updated_at"])


def _check_site_threadsafe(site: Site, *, check_type: str, performed_by) -> dict:
    try:
        return {
            "id": site.id,
            **test_connection(
                site, check_type=check_type, performed_by=performed_by
            ),
        }
    finally:
        # Each worker thread opens its own (thread-local) DB connection; close it
        # so the pool does not leak connections every health-check round.
        connection.close()


# Grace (s) subtracted from the due cutoffs so beat/check-execution jitter does
# not bump a site to the next tick (e.g. a site checked ~599s ago still counts as
# due for its 600s window). Small relative to the gap between the two intervals.
_DUE_TOLERANCE_SECONDS = 30


def _due_filter():
    """Q for sites whose next periodic check is due now.

    A site that passed its last check (status UP) is re-checked every
    ``SITE_HEALTHCHECK_OK_INTERVAL_SECONDS``; a site that failed or was never
    reached (DOWN/UNKNOWN, or never checked) is retried every
    ``SITE_HEALTHCHECK_FAIL_INTERVAL_SECONDS``. This keeps healthy sites off the
    every-tick cadence so we don't hammer them needlessly.
    """
    now = timezone.now()
    ok_cutoff = now - timedelta(
        seconds=settings.SITE_HEALTHCHECK_OK_INTERVAL_SECONDS - _DUE_TOLERANCE_SECONDS
    )
    fail_cutoff = now - timedelta(
        seconds=settings.SITE_HEALTHCHECK_FAIL_INTERVAL_SECONDS - _DUE_TOLERANCE_SECONDS
    )
    return (
        Q(last_checked_at__isnull=True)
        | Q(status=Site.Status.UP, last_checked_at__lte=ok_cutoff)
        | (~Q(status=Site.Status.UP) & Q(last_checked_at__lte=fail_cutoff))
    )


def check_hosting(
    hosting_id, *, check_type: str = "periodic", performed_by=None
) -> list[dict]:
    """Health-check the due sites of one hosting, at most ``check_concurrency``
    domains at a time (the rest queue and run as slots free up). Different
    hostings are checked in parallel by the Celery fan-out in
    apps/monitoring/tasks.py. ``hosting_id=None`` checks sites with no hosting.

    Defaults to ``check_type="periodic"`` (the Celery caller), which only checks
    sites that are *due* (see ``_due_filter`` — healthy sites are re-checked less
    often than failed ones). The on-demand "Check ngay" button passes ``manual``
    + the acting user and checks every site regardless of when it was last seen.
    """
    qs = Site.objects.filter(hosting_id=hosting_id, is_deleted=False)
    if check_type == "periodic":
        qs = qs.filter(_due_filter())
    # Primary sites ("trang chính") are submitted to the pool first so each
    # round checks them before the rest of the hosting's sites.
    sites = list(qs.order_by("-is_primary", "id"))
    if not sites:
        return []

    concurrency = DEFAULT_CHECK_CONCURRENCY
    if hosting_id is not None:
        hosting = Hosting.objects.filter(id=hosting_id, is_deleted=False).first()
        if hosting:
            concurrency = max(1, hosting.check_concurrency)

    worker = partial(
        _check_site_threadsafe, check_type=check_type, performed_by=performed_by
    )
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        return list(executor.map(worker, sites))


def import_sites_from_xlsx(file, hosting=None) -> dict:
    """Parse an uploaded .xlsx and bulk-create sites, optionally assigning every
    created site to ``hosting``.

    Expected header row: name, base_url, consumer_key, consumer_secret, plus an
    optional ``platform`` column (woocommerce | sapo; defaults to woocommerce).
    Returns ``{"created": int, "errors": [{"row": int, "error": str}]}``.
    Rows missing required data or whose base_url already exists are skipped
    with an error entry (so partial imports surface clearly).
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return {"created": 0, "errors": [{"row": 0, "error": "File không đọc được (.xlsx)."}]}

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {"created": 0, "errors": [{"row": 0, "error": "File rỗng."}]}

    cols = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
    missing = [c for c in IMPORT_COLUMNS if c not in cols]
    if missing:
        return {"created": 0, "errors": [{"row": 1, "error": f"Thiếu cột: {', '.join(missing)}"}]}

    created = 0
    errors: list[dict] = []
    for idx, row in enumerate(rows, start=2):
        data = {}
        for col in IMPORT_COLUMNS:
            i = cols[col]
            value = row[i] if i < len(row) else None
            data[col] = str(value).strip() if value is not None else ""

        if not all(data.values()):
            errors.append({"row": idx, "error": "Thiếu dữ liệu bắt buộc."})
            continue
        if Site.objects.filter(base_url=data["base_url"], is_deleted=False).exists():
            errors.append({"row": idx, "error": f"base_url đã tồn tại: {data['base_url']}"})
            continue

        platform = Site.Platform.WOOCOMMERCE
        if "platform" in cols:
            i = cols["platform"]
            raw = row[i] if i < len(row) else None
            value = str(raw).strip().lower() if raw is not None else ""
            if value:
                if value not in Site.Platform.values:
                    errors.append({"row": idx, "error": f"platform không hợp lệ: {value}"})
                    continue
                platform = value

        create_site(**data, hosting=hosting, platform=platform)
        created += 1

    return {"created": created, "errors": errors}


def _parse_check_concurrency(value) -> int:
    """Coerce a cell value to a valid check_concurrency (1..MAX), default on junk."""
    try:
        n = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return DEFAULT_CHECK_CONCURRENCY
    if n <= 0:
        return DEFAULT_CHECK_CONCURRENCY
    return min(MAX_CHECK_CONCURRENCY, n)


def import_hostings_from_xlsx(file) -> dict:
    """Parse an uploaded .xlsx and bulk-create hostings.

    Expected header row: name (required), provider, account_username, note,
    check_concurrency (optional; defaults to 5 when blank/invalid, clamped 1..50).
    Returns ``{"created": int, "errors": [{"row": int, "error": str}]}``.
    Rows missing a name or duplicating an existing (name, account_username) are
    skipped with an error entry (so partial imports surface clearly).
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return {"created": 0, "errors": [{"row": 0, "error": "File không đọc được (.xlsx)."}]}

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {"created": 0, "errors": [{"row": 0, "error": "File rỗng."}]}

    cols = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
    missing = [c for c in HOSTING_REQUIRED_COLUMNS if c not in cols]
    if missing:
        return {"created": 0, "errors": [{"row": 1, "error": f"Thiếu cột: {', '.join(missing)}"}]}

    def cell(row, col):
        i = cols.get(col)
        if i is None or i >= len(row):
            return ""
        value = row[i]
        return str(value).strip() if value is not None else ""

    created = 0
    errors: list[dict] = []
    for idx, row in enumerate(rows, start=2):
        name = cell(row, "name")
        if not name:
            errors.append({"row": idx, "error": "Thiếu tên hosting."})
            continue

        account = cell(row, "account_username")
        if Hosting.objects.filter(
            name=name, account_username=account, is_deleted=False
        ).exists():
            errors.append({"row": idx, "error": f"Hosting đã tồn tại: {name}"})
            continue

        create_hosting(
            name=name,
            provider=cell(row, "provider"),
            account_username=account,
            note=cell(row, "note"),
            check_concurrency=_parse_check_concurrency(cell(row, "check_concurrency")),
        )
        created += 1

    return {"created": created, "errors": errors}


# --- Site notes --------------------------------------------------------------

# Rich-text allowlist matching what the TipTap editor can produce. Anything
# else (scripts, event handlers, inline styles, images) is stripped — note
# content is HTML stored verbatim, so it must be sanitized before saving.
NOTE_ALLOWED_TAGS = {
    "p", "br", "strong", "b", "em", "i", "u", "s", "strike", "del",
    "h1", "h2", "h3", "ul", "ol", "li", "blockquote", "code", "pre", "a",
}
# ``rel`` is intentionally omitted: nh3/ammonia manages it via ``link_rel``
# below (setting both panics).
NOTE_ALLOWED_ATTRIBUTES = {"a": {"href", "title", "target"}}

# Image attachment limits.
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}
MAX_IMAGE_BYTES = 5 * 1024 * 1024  # 5 MB per attachment


def sanitize_note_html(html: str) -> str:
    """Strip unsafe markup from note HTML, keeping only the rich-text tags the
    editor produces. Prevents stored XSS between admin users."""
    if not html:
        return ""
    return nh3.clean(
        html,
        tags=NOTE_ALLOWED_TAGS,
        attributes=NOTE_ALLOWED_ATTRIBUTES,
        link_rel="noopener noreferrer nofollow",
    )


def _validate_image(upload) -> None:
    """Reject attachments that are not small images (type + size)."""
    content_type = getattr(upload, "content_type", "") or ""
    if content_type not in ALLOWED_IMAGE_TYPES:
        raise ValidationError(
            {"images": f"Định dạng ảnh không hỗ trợ: {content_type or 'không rõ'}."}
        )
    if upload.size and upload.size > MAX_IMAGE_BYTES:
        raise ValidationError({"images": "Ảnh vượt quá 5MB."})


def _attach_images(note: SiteNote, images) -> None:
    for upload in images:
        _validate_image(upload)
        SiteNoteImage.objects.create(
            note=note,
            image=upload,
            original_name=(upload.name or "")[:255],
        )


@transaction.atomic
def create_site_note(*, site, content, created_by=None, images=()) -> SiteNote:
    note = SiteNote.objects.create(
        site=site,
        content=sanitize_note_html(content or ""),
        created_by=created_by,
    )
    _attach_images(note, images)
    return note


@transaction.atomic
def update_site_note(
    note: SiteNote, *, content=None, new_images=(), remove_image_ids=()
) -> SiteNote:
    if content is not None:
        note.content = sanitize_note_html(content)
        note.save(update_fields=["content", "updated_at"])
    if remove_image_ids:
        # Scope the delete to this note so a caller can't remove another's image.
        note.images.filter(id__in=remove_image_ids).delete()
    _attach_images(note, new_images)
    return note


def delete_site_note(note: SiteNote) -> None:
    """Soft-delete a note (kept in the DB for history/audit, hidden from list)."""
    note.is_deleted = True
    note.deleted_at = timezone.now()
    note.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
